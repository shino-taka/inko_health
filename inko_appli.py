#インコの健康管理

import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op
from openpyxl.styles import Alignment
import os # ファイルの存在チェックのために使用
import sys
import datetime as dt
import time

excel_file = 'inko_health.xlsx'

# Excelファイルが存在しない場合に新規作成する
if not os.path.exists(excel_file):
    try:
      wb = op.Workbook()
      wb.save(excel_file)
      messagebox.showinfo("ファイル作成", f"'{excel_file}' を新規作成しました。")
      time.sleep(0.1) #ファイル作成後に少し待つ
    except Exception as e:
      print(f"ファイル作成エラー:{e}")
      messagebox.showerror("ファイル作成エラー",f"ファイルの新規作成に失敗しました\n詳細: {e}")
      sys.exit()


def clear():
  """入力データを消す"""
  name.set("")
  # data.delete(0,tk.END)
  weigth.delete(0,tk.END)
  food.delete(0,tk.END)
  drug.delete("1.0",tk.END)
  condition.delete("1.0",tk.END)
  etc.delete("1.0",tk.END)

def save_to_excel(event):
  """入力されたデータをエクセルに保存する関数"""
  get_name = name.get()
  get_data = data.get()
  get_weigth = weigth.get()
  get_food = food.get()
  get_drug = drug.get("1.0", tk.END).strip()
  get_con = condition.get("1.0", tk.END).strip()
  get_etc = etc.get("1.0", tk.END).strip()
  
  try:
    # Excelファイルを読み込む
    wb = op.load_workbook(excel_file)
  except Exception as e:
    messagebox.showerror("エラー", f"エクセルファイルの読み込み中にエラーが発生しました\n{e}\nファイルが破損している可能性があります。")
    return

  # 既存のシート名リストを取得
  sheet_names = wb.sheetnames
  # 選択された名前をシート名として使用
  sheet_name = get_name

  if sheet_name in sheet_names:
    # 既存のシートがあればそれを取得
    ws = wb[sheet_name]
  else:
    # 既存のシートがなければ新規作成
    ws = wb.create_sheet(sheet_name)
    # 新規シートの場合、ヘッダーを追加 (一度だけ)
    # if ws.max_row == 0 or ws.cell(row=1, column=1).value is None: # シートが完全に空の場合
    ws.append(["日付", "体重", "食事量","薬","体調","その他"])

  #データをシートに追加
  ws.append([get_data, get_weigth, get_food, get_drug, get_con, get_etc])
  #textウィジェットの改行をそのまま表示(セルの書式設定を折り返し表示にする)
  last_row_index = ws.max_row 
  ws.cell(row=last_row_index, column=4).alignment = Alignment(wrapText=True) # 薬のセル
  ws.cell(row=last_row_index, column=5).alignment = Alignment(wrapText=True) # 体調のセル
  ws.cell(row=last_row_index, column=6).alignment = Alignment(wrapText=True) # その他のセル
  # 変更を保存
  wb.save(excel_file)
  messagebox.showinfo("成功", f"データがシート '{sheet_name}' に書き込まれました。")
  clear() # 入力欄をクリア

#GUI画面
root = tk.Tk()
root.title("インコ体調管理")
root.minsize(400,300)

name_label = tk.Label(text="名前")
name_label.grid(row=0, column=0, sticky=tk.E,padx=10, pady=10)
name = ttk.Combobox()
name.grid(row=0, column=1, sticky=tk.W,padx=10, pady=10)
name["value"] = ("ブラン","エルモ","まめのすけ","ルビー")
name.set("")

data_label = tk.Label(text="日付")
data_label.grid(row=2, column=0, sticky=tk.E, padx=5, pady=5)
data = tk.Entry()
data.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
#今日の日付を自動入力
today = dt.date.today().strftime("%Y/%m/%d")
data.insert(0, today)

weigth_label = tk.Label(text="体重")
weigth_label.grid(row=3, column=0, sticky=tk.E, padx=5, pady=5)
weigth = tk.Entry()
weigth.grid(row=3, column=1, sticky=tk.W, padx=5, pady=5)

food_label = tk.Label(text="食事量")
food_label.grid(row=4, column=0, sticky=tk.E, padx=5, pady=5)
food = tk.Entry()
food.grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)

drug_label = tk.Label(text="薬")
drug_label.grid(row=7, column=0, sticky=tk.E, padx=5, pady=5)
drug = tk.Text(width=40,height=3)
drug.grid(row=7, column=1, sticky=tk.W, padx=5, pady=5)

condition_label = tk.Label(text="体調")
condition_label.grid(row=10, column=0, sticky=tk.E, padx=5, pady=5)
condition = tk.Text(width=40,height=3)
condition.grid(row=10, column=1, sticky=tk.W, padx=5, pady=5)

etc_label = tk.Label(text="その他")
etc_label.grid(row=13, column=0, sticky=tk.E, padx=5, pady=5)
etc = tk.Text(width=40,height=3)
etc.grid(row=13, column=1, sticky=tk.W, padx=5, pady=5)

btn = tk.Button(text="登録する")
btn.grid(row=14, column=0, columnspan=2,sticky=tk.NSEW,padx=10, pady=10)
btn.bind("<1>",save_to_excel)

#中央寄せ(ウィンドウサイズ変更に対応可)
#列の伸縮設定
root.grid_columnconfigure(0, weight=1) # 0列目 (ラベル側)
root.grid_columnconfigure(1, weight=3) # 1列目 (入力欄側) - こちらをより広げる

#行の伸縮設定
for i in range(15): # 必要に応じて列数を増やす
  root.grid_rowconfigure(i, weight=1)

root.mainloop()